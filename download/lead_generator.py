#!/usr/bin/env python3
"""
Lead Generation Script — Google Maps Business Scraper
=====================================================
Searches real businesses via Google Maps API, enriches each lead
with website-age and mobile-friendliness checks, scores them
(HOT / WARM / COLD), and exports colour-coded Excel + CSV files.

Requirements
------------
    pip install googlemaps requests beautifulsoup4 openpyxl colorama
"""

# ╔══════════════════════════════════════════════════════════════════╗
# ║                        CONFIGURATION                           ║
# ╚══════════════════════════════════════════════════════════════════╝

GOOGLE_API_KEY: str = "YOUR_GOOGLE_API_KEY_HERE"  # <-- replace with your key

CITY: str = "New York"                             # Target city
SEARCH_QUERY: str = "plumber"                     # Niche / keyword
MAX_RESULTS: int = 20                             # Businesses to retrieve

# Scoring thresholds (based on estimated website year)
HOT_YEAR_THRESHOLD = 2018      # <= this year → HOT
WARM_YEAR_UPPER = 2021         # <= this year and > HOT → WARM
# > WARM_YEAR_UPPER → COLD

# Requests & scraping
REQUEST_TIMEOUT = 10           # seconds
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0.0.0 Safari/537.36"
)

# Output file paths
EXCEL_OUTPUT = "leads_output.xlsx"
CSV_OUTPUT = "leads_output.csv"

# ╔══════════════════════════════════════════════════════════════════╗
# ║                        IMPORTS                                 ║
# ╚══════════════════════════════════════════════════════════════════╝

import csv
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

import googlemaps
import requests
from bs4 import BeautifulSoup
from colorama import Fore, Style, init
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Colourama auto-reset
init(autoreset=True)

# ╔══════════════════════════════════════════════════════════════════╗
# ║                     CONSTANTS & HELPERS                        ║
# ╚══════════════════════════════════════════════════════════════════╝

# Excel colour fills for lead scoring
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_GREEN = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

FONT_BOLD_WHITE = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD_BLACK = Font(bold=True, color="000000", size=11)
FONT_HEADER = Font(bold=True, color="FFFFFF", size=12)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class Logger:
    """Minimal terminal logger with progress counter."""

    def __init__(self, total: int):
        self.total = total
        self.current = 0
        self.errors = 0

    def _progress(self) -> str:
        pct = (self.current / self.total * 100) if self.total else 0
        return f"[{self.current}/{self.total}] {pct:5.1f}%"

    def tick(self, name: str) -> None:
        self.current += 1
        print(f"  {Fore.CYAN}{self._progress()}{Style.RESET_ALL}  Processing: {name}")

    def warn(self, msg: str) -> None:
        self.errors += 1
        print(f"  {Fore.YELLOW}⚠  WARNING:{Style.RESET_ALL} {msg}")

    def err(self, msg: str) -> None:
        self.errors += 1
        print(f"  {Fore.RED}✖  ERROR:{Style.RESET_ALL}   {msg}")

    def ok(self, msg: str) -> None:
        print(f"  {Fore.GREEN}✔  {msg}{Style.RESET_ALL}")


# ╔══════════════════════════════════════════════════════════════════╗
# ║                     CORE FUNCTIONS                             ║
# ╚══════════════════════════════════════════════════════════════════╝


def fetch_businesses(
    gmaps: googlemaps.Client,
    query: str,
    city: str,
    max_results: int,
    log: Logger,
) -> list[dict]:
    """Search Google Maps for businesses matching *query* in *city*."""
    full_query = f"{query} in {city}"
    print(f"\n{Fore.BLUE}🔍 Searching: \"{full_query}\" (max {max_results}){Style.RESET_ALL}\n")

    try:
        places_result = gmaps.places(query=full_query)
    except googlemaps.exceptions.ApiError as exc:
        print(f"{Fore.RED}Google Maps API error: {exc}{Style.RESET_ALL}")
        sys.exit(1)
    except Exception as exc:
        print(f"{Fore.RED}Unexpected API error: {exc}{Style.RESET_ALL}")
        sys.exit(1)

    raw_results = places_result.get("results", [])
    if not raw_results:
        print(f"{Fore.YELLOW}No results found for \"{full_query}\".{Style.RESET_ALL}")
        return []

    businesses: list[dict] = []
    seen_ids: set[str] = set()

    for place in raw_results[:max_results]:
        place_id = place.get("place_id")
        if place_id in seen_ids:
            continue
        seen_ids.add(place_id)

        name = place.get("name", "N/A")
        log.tick(name)

        # ---- gather fields from the summary response ----
        phone = "N/A"
        website = None
        address = place.get("formatted_address", "N/A")
        rating = place.get("rating", None)
        reviews = place.get("user_ratings_total", 0)

        # ---- details request (phone + website) ----
        try:
            details = gmaps.place(place_id=place_id, fields=["formatted_phone_number", "website"])
            detail = details.get("result", {})
            phone = detail.get("formatted_phone_number") or phone
            website = detail.get("website")  # may be None
        except Exception as exc:
            log.err(f"Details fetch failed for \"{name}\": {exc}")

        businesses.append({
            "name": name,
            "phone": phone,
            "website": website if website else "N/A",
            "address": address,
            "rating": rating,
            "reviews": reviews,
        })

        # Be gentle with the API (free tier: 50 req/s, but let's be safe)
        time.sleep(0.15)

    return businesses


def analyze_website(url: str, log: Logger, business_name: str) -> dict:
    """
    Fetch a website and extract:
      - copyright_year (best-effort integer or None)
      - is_mobile_friendly (bool — checks for <meta name=\"viewport\">)
    """
    result = {"copyright_year": None, "is_mobile_friendly": False}

    if not url or url == "N/A":
        return result

    # Ensure scheme is present
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    try:
        resp = requests.get(
            url,
            headers={"User-Agent": USER_AGENT},
            timeout=REQUEST_TIMEOUT,
            allow_redirects=True,
        )
        resp.raise_for_status()
    except requests.exceptions.SSLError:
        log.warn(f"SSL error for {business_name}, trying HTTP …")
        try:
            resp = requests.get(
                url.replace("https://", "http://"),
                headers={"User-Agent": USER_AGENT},
                timeout=REQUEST_TIMEOUT,
                allow_redirects=True,
            )
            resp.raise_for_status()
        except Exception as inner_exc:
            log.err(f"Could not reach website of \"{business_name}\": {inner_exc}")
            return result
    except Exception as exc:
        log.err(f"Could not reach website of \"{business_name}\": {exc}")
        return result

    content_type = resp.headers.get("Content-Type", "")
    if "text/html" not in content_type:
        log.warn(f"Non-HTML content at {business_name} ({content_type})")
        return result

    try:
        soup = BeautifulSoup(resp.text, "html.parser")
    except Exception:
        log.err(f"HTML parse error for {business_name}")
        return result

    # --- Copyright year detection ---
    year = _extract_copyright_year(soup, resp.text)
    result["copyright_year"] = year

    # --- Mobile-friendly check (viewport meta tag) ---
    viewport = soup.find("meta", attrs={"name": re.compile(r"^viewport$", re.I)})
    if viewport and viewport.get("content"):
        result["is_mobile_friendly"] = True

    return result


def _extract_copyright_year(soup: BeautifulSoup, text: str) -> Optional[int]:
    """
    Attempt to find the copyright / site-creation year.

    Strategy (first match wins):
      1. <meta name="copyright" content="… YYYY …">
      2. Text matching  © … YYYY  or  ™ … YYYY
      3. Any "©" line containing a 4-digit year
      4. Fallback: look for the oldest year in footer-like elements
    """
    # 1. Meta copyright tag
    meta = soup.find("meta", attrs={"name": re.compile(r"copyright", re.I)})
    if meta and meta.get("content"):
        years = re.findall(r"\b(19|20)\d{2}\b", meta["content"])
        if years:
            return int(years[-1])

    # 2–3. © or ™ with year anywhere in the page
    #    Prioritise footer blocks
    footer = soup.find("footer")
    search_text = footer.get_text() if footer else text

    # Look for the explicit "© YYYY" pattern first
    explicit = re.findall(
        r"(?:©|&copy;|\(c\)|™)\s*(?:\d{4}\s*[-–]\s*)?(\d{4})", search_text
    )
    if explicit:
        # Return the earliest year (likely site creation)
        return int(min(explicit))

    # 4. Broader search — any year near the word "copyright"
    broad = re.findall(r"(?:copyright|©|&copy;)[^0-9]*((?:19|20)\d{2})", search_text, re.I)
    if broad:
        return int(min(broad))

    # 5. Last resort: look at the footer for any 4-digit year
    if footer:
        footer_years = re.findall(r"\b(19|20)\d{2}\b", footer.get_text())
        if footer_years:
            return int(footer_years[0])

    return None


def score_lead(website: str, copyright_year: Optional[int], is_mobile_friendly: bool) -> str:
    """
    Assign a temperature score:
      HOT  — no website OR site estimated before HOT_YEAR_THRESHOLD
      WARM — site between HOT_YEAR_THRESHOLD+1 and WARM_YEAR_UPPER
      COLD — modern site (after WARM_YEAR_UPPER)
    """
    if website == "N/A":
        return "HOT"

    if copyright_year is None:
        # Has a website but we couldn't determine the year — assume WARM
        return "WARM"

    if copyright_year <= HOT_YEAR_THRESHOLD:
        return "HOT"
    elif copyright_year <= WARM_YEAR_UPPER:
        return "WARM"
    else:
        return "COLD"


def save_to_excel(leads: list[dict], filepath: str) -> None:
    """Write leads to a colour-coded .xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = [
        "#",
        "Name",
        "Phone",
        "Website",
        "Address",
        "Rating",
        "Reviews",
        "Site Year",
        "Mobile Friendly",
        "Score",
    ]

    # Header row
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, lead in enumerate(leads, 2):
        score = lead["score"]
        if score == "HOT":
            row_fill, row_font = FILL_RED, FONT_BOLD_WHITE
        elif score == "WARM":
            row_fill, row_font = FILL_YELLOW, FONT_BOLD_BLACK
        else:
            row_fill, row_font = FILL_GREEN, FONT_BOLD_WHITE

        values = [
            row_idx - 1,
            lead["name"],
            lead["phone"],
            lead["website"],
            lead["address"],
            lead["rating"] if lead["rating"] else "N/A",
            lead["reviews"],
            lead["copyright_year"] if lead["copyright_year"] else "N/A",
            "Yes" if lead["is_mobile_friendly"] else "No",
            lead["score"],
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = row_font
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Column widths
    widths = [5, 28, 18, 36, 36, 8, 10, 12, 16, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)


def save_to_csv(leads: list[dict], filepath: str) -> None:
    """Write leads to a .csv file."""
    fieldnames = [
        "name",
        "phone",
        "website",
        "address",
        "rating",
        "reviews",
        "copyright_year",
        "is_mobile_friendly",
        "score",
    ]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for lead in leads:
            row = {
                "name": lead["name"],
                "phone": lead["phone"],
                "website": lead["website"],
                "address": lead["address"],
                "rating": lead["rating"] if lead["rating"] else "",
                "reviews": lead["reviews"],
                "copyright_year": lead["copyright_year"] if lead["copyright_year"] else "",
                "is_mobile_friendly": lead["is_mobile_friendly"],
                "score": lead["score"],
            }
            writer.writerow(row)


def print_summary(leads: list[dict], elapsed: float) -> None:
    """Print a coloured summary table to the terminal."""
    hot = [l for l in leads if l["score"] == "HOT"]
    warm = [l for l in leads if l["score"] == "WARM"]
    cold = [l for l in leads if l["score"] == "COLD"]
    total = len(leads)

    mobile_yes = sum(1 for l in leads if l["is_mobile_friendly"])
    mobile_no = total - mobile_yes
    no_website = sum(1 for l in leads if l["website"] == "N/A")

    sep = "═" * 62

    print(f"\n{sep}")
    print(f"  {Fore.WHITE + Style.BRIGHT}LEAD GENERATION SUMMARY{Style.RESET_ALL}")
    print(sep)
    print(f"  Date         : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  City         : {CITY}")
    print(f"  Niche        : {SEARCH_QUERY}")
    print(f"  Elapsed      : {elapsed:.1f}s")
    print(sep)
    print(f"  Total leads  : {total}")
    print(f"  {Fore.RED + Style.BRIGHT}HOT  leads  : {len(hot)}{Style.RESET_ALL}   (no site or site ≤ {HOT_YEAR_THRESHOLD})")
    print(f"  {Fore.YELLOW + Style.BRIGHT}WARM leads  : {len(warm)}{Style.RESET_ALL}   (site {HOT_YEAR_THRESHOLD+1}–{WARM_YEAR_UPPER})")
    print(f"  {Fore.GREEN + Style.BRIGHT}COLD leads  : {len(cold)}{Style.RESET_ALL}   (site > {WARM_YEAR_UPPER})")
    print(sep)
    print(f"  No website   : {no_website}")
    print(f"  Mobile-Friendly : {mobile_yes}   |   Not Mobile-Friendly : {mobile_no}")
    print(sep)

    # Top-rated businesses
    rated = sorted(
        [l for l in leads if l["rating"]],
        key=lambda x: (x["rating"], x["reviews"]),
        reverse=True,
    )
    if rated:
        print(f"\n  {Fore.CYAN}Top Rated Businesses:{Style.RESET_ALL}")
        for i, b in enumerate(rated[:5], 1):
            print(
                f"    {i}. {b['name']}"
                f"  ★ {b['rating']}  ({b['reviews']} reviews)"
                f"  {Fore.MAGENTA}[{b['score']}]{Style.RESET_ALL}"
            )

    # HOT leads list (best prospects)
    if hot:
        print(f"\n  {Fore.RED + Style.BRIGHT}🔥 HOT Leads (best prospects):{Style.RESET_ALL}")
        for i, b in enumerate(hot, 1):
            reason = "No website" if b["website"] == "N/A" else f"Site ~{b['copyright_year']}"
            print(f"    {i}. {b['name']}  —  {reason}  —  {b['phone']}")

    print(f"\n{sep}")
    print(f"  {Fore.GREEN}Excel → {Path(EXCEL_OUTPUT).resolve()}{Style.RESET_ALL}")
    print(f"  {Fore.GREEN}CSV   → {Path(CSV_OUTPUT).resolve()}{Style.RESET_ALL}")
    print(sep)


# ╔══════════════════════════════════════════════════════════════════╗
# ║                         MAIN                                   ║
# ╚══════════════════════════════════════════════════════════════════╝


def main() -> None:
    banner = f"""
╔══════════════════════════════════════════════════════════╗
║        🎯  Google Maps Lead Generator  🎯               ║
╚══════════════════════════════════════════════════════════╝"""
    print(Fore.CYAN + banner + Style.RESET_ALL)

    # ── Validate API key ─────────────────────────────────────
    if not GOOGLE_API_KEY or GOOGLE_API_KEY == "YOUR_GOOGLE_API_KEY_HERE":
        print(
            f"\n{Fore.RED}✖  ERROR: Please set your GOOGLE_API_KEY at the top of the script.{Style.RESET_ALL}\n"
        )
        sys.exit(1)

    # ── Initialise ───────────────────────────────────────────
    start_time = time.time()
    gmaps = googlemaps.Client(key=GOOGLE_API_KEY)
    log = Logger(total=MAX_RESULTS)

    # ── Step 1: Fetch businesses from Google Maps ────────────
    print(f"\n{Fore.MAGENTA}── Step 1/3 ─ Fetching businesses from Google Maps ──{Style.RESET_ALL}")
    businesses = fetch_businesses(gmaps, SEARCH_QUERY, CITY, MAX_RESULTS, log)

    if not businesses:
        print(f"\n{Fore.YELLOW}No businesses found. Exiting.{Style.RESET_ALL}")
        sys.exit(0)

    print(f"\n  {Fore.GREEN}✔ Found {len(businesses)} businesses.{Style.RESET_ALL}")

    # ── Step 2: Analyse websites ─────────────────────────────
    print(f"\n{Fore.MAGENTA}── Step 2/3 ─ Analysing websites ──{Style.RESET_ALL}\n")
    for biz in businesses:
        analysis = analyze_website(biz["website"], log, biz["name"])
        biz["copyright_year"] = analysis["copyright_year"]
        biz["is_mobile_friendly"] = analysis["is_mobile_friendly"]
        biz["score"] = score_lead(
            biz["website"], analysis["copyright_year"], analysis["is_mobile_friendly"]
        )
        time.sleep(0.3)  # polite delay

    # ── Step 3: Save results ─────────────────────────────────
    print(f"\n{Fore.MAGENTA}── Step 3/3 ─ Saving results ──{Style.RESET_ALL}")

    save_to_excel(businesses, EXCEL_OUTPUT)
    print(f"  {Fore.GREEN}✔ Excel saved: {EXCEL_OUTPUT}{Style.RESET_ALL}")

    save_to_csv(businesses, CSV_OUTPUT)
    print(f"  {Fore.GREEN}✔ CSV saved:   {CSV_OUTPUT}{Style.RESET_ALL}")

    # ── Summary ──────────────────────────────────────────────
    elapsed = time.time() - start_time
    print_summary(businesses, elapsed)


if __name__ == "__main__":
    main()
